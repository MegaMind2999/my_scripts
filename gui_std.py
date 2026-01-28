import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import urllib3
import time
from datetime import datetime
import sys
import base64
import json

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font

# =========================================================
# CHROME RENDERING IMPORTS
# =========================================================
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =========================================================
# CONFIGURATION
# =========================================================
COLOR_BG = "#f0f0f0"
COLOR_CARD = "#ffffff"
COLOR_HEADER = "#1a1d29"
COLOR_ACCENT = "#6366f1"
COLOR_ACCENT_HOVER = "#4f46e5"
COLOR_TEXT = "#1f2937"
COLOR_TEXT_LIGHT = "#6b7280"
COLOR_SUCCESS = "#10b981"
COLOR_WARNING = "#ef4444"
COLOR_BATCH = "#8b5cf6"
COLOR_BORDER = "#e5e7eb"
COLOR_SHADOW = "#00000010"

LOGIN_URL = "https://tdb.tanta.edu.eg/student_results/default.aspx"
TARGET_URL = "https://tdb.tanta.edu.eg/student_results/marklist.aspx"
REPORT_URL = "https://tdb.tanta.edu.eg/student_results/marklist_report.aspx"

ACCOUNTS = {
    "Science": {
        "user": "علوم",
        "pass": "علوم1234",
        "label": "👤 Science"
    },
    "Education": {
        "user": "نهال شبل",
        "pass": "الحمدلله12345",
        "label": "👤 Education"
    }
}

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class TantaScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Student List Manager - Chrome Portrait")
        
        icon_path = resource_path("nicde.ico")
        try:
            self.root.iconbitmap(icon_path)
        except Exception:
            pass

        self.root.geometry("900x1000")
        self.root.configure(bg=COLOR_BG)
        
        self.current_profile = "Science"
        self.stop_event = threading.Event()
        self.is_batch_running = False
        
        self.var_save_excel = tk.BooleanVar(value=True)
        self.var_save_pdf = tk.BooleanVar(value=False)
        self.var_cond_format = tk.BooleanVar(value=False)
        
        self.session = None
        self.reset_session()
        
        self.current_soup = None
        self.selections = {}
        self.dropdown_map = {}
        self.combos = {}
        
        self.id_map = {
            "Year": "ctl00_ContentPlaceHolder3_ddl_acad_year",
            "Faculty": "ctl00_ContentPlaceHolder3_ddl_fac",
            "Regulation": "ctl00_ContentPlaceHolder3_ddl_bylaw",
            "Phase": "ctl00_ContentPlaceHolder3_ddl_phase",
            "Dept": "ctl00_ContentPlaceHolder3_ddl_dept",
            "Semester": "ctl00_ContentPlaceHolder3_ddl_semester",
            "Door": "ctl00_ContentPlaceHolder3_door",
            "SubjSem": "ctl00_ContentPlaceHolder3_ddl_semester_subject",
            "Subject": "ctl00_ContentPlaceHolder3_ddl_subj"
        }
        
        self._apply_styles()
        self._setup_ui()
        
        self.log("🚀 System Initialized", color="info")
        threading.Thread(target=self.perform_login, daemon=True).start()

    def reset_session(self):
        self.session = requests.Session()
        retry_strategy = Retry(
            total=4, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Origin': 'https://tdb.tanta.edu.eg',
            'Referer': LOGIN_URL
        })

    def _apply_styles(self):
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except:
            pass
        
        style.configure("Card.TFrame", background=COLOR_CARD, relief="flat")
        style.configure("Main.TFrame", background=COLOR_BG)
        style.configure("Modern.TLabel", background=COLOR_CARD, foreground=COLOR_TEXT, 
                       font=("Segoe UI", 10))
        style.configure("ModernBold.TLabel", background=COLOR_CARD, foreground=COLOR_TEXT, 
                       font=("Segoe UI", 10, "bold"))
        
        style.configure("Modern.TCombobox", 
                       fieldbackground=COLOR_CARD,
                       background=COLOR_CARD,
                       borderwidth=1,
                       relief="flat")
        style.map('Modern.TCombobox',
                 fieldbackground=[('readonly', COLOR_CARD)],
                 selectbackground=[('readonly', COLOR_ACCENT)])

    def _setup_ui(self):
        header_frame = tk.Frame(self.root, bg=COLOR_HEADER, height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        header_content = tk.Frame(header_frame, bg=COLOR_HEADER)
        header_content.place(relx=0.5, rely=0.5, anchor='center')
        
        tk.Label(header_content, text="📊", bg=COLOR_HEADER, fg="white", 
                font=("Segoe UI", 32)).pack(side=tk.LEFT, padx=(0, 15))
        
        title_frame = tk.Frame(header_content, bg=COLOR_HEADER)
        title_frame.pack(side=tk.LEFT)
        
        tk.Label(title_frame, text="Student List Manager", bg=COLOR_HEADER, fg="white",
                font=("Segoe UI", 20, "bold")).pack(anchor='w')
        tk.Label(title_frame, text="Export ro Excel/PDF", 
                bg=COLOR_HEADER, fg="#9ca3af", font=("Segoe UI", 9)).pack(anchor='w')
        
        profile_btn_frame = tk.Frame(header_frame, bg=COLOR_HEADER)
        profile_btn_frame.pack(side=tk.RIGHT, padx=20)
        
        self.btn_profile = tk.Button(
            profile_btn_frame,
            text=ACCOUNTS[self.current_profile]['label'],
            command=self.toggle_account,
            bg="#374151", fg="white", activebackground="#4b5563",
            font=("Segoe UI", 10, "bold"), bd=0, padx=20, pady=10,
            cursor="hand2", relief=tk.FLAT
        )
        self.btn_profile.pack()
        
        scroll_container = tk.Frame(self.root, bg=COLOR_BG)
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(scroll_container, bg=COLOR_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        
        main_container = tk.Frame(canvas, bg=COLOR_BG)
        canvas_frame = canvas.create_window((0, 0), window=main_container, anchor="nw")
        
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        def configure_canvas_width(event):
            canvas.itemconfig(canvas_frame, width=event.width)
        
        main_container.bind("<Configure>", configure_scroll_region)
        canvas.bind("<Configure>", configure_canvas_width)
        
        def on_mousewheel(event):
            widget = self.root.winfo_containing(event.x_root, event.y_root)
            is_over_combobox = False
            while widget:
                if isinstance(widget, ttk.Combobox) or widget.winfo_class() == 'Toplevel':
                    is_over_combobox = True
                    break
                widget = widget.master

            if not is_over_combobox:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        self.root.bind("<MouseWheel>", on_mousewheel)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        main_container.pack_configure = lambda **kw: None
        tk.Frame(main_container, bg=COLOR_BG, height=30).grid(row=0, column=0)
        content_frame = tk.Frame(main_container, bg=COLOR_BG)
        content_frame.grid(row=1, column=0, padx=30, sticky="ew")
        main_container.grid_columnconfigure(0, weight=1)
        
        main_container = content_frame
        
        selection_card = tk.Frame(main_container, bg=COLOR_CARD, relief=tk.FLAT, bd=0)
        selection_card.pack(fill=tk.X, pady=(0, 20))
        
        card_header = tk.Frame(selection_card, bg=COLOR_ACCENT, height=50)
        card_header.pack(fill=tk.X)
        card_header.pack_propagate(False)
        
        tk.Label(card_header, text="📋 Selection Criteria", bg=COLOR_ACCENT, fg="white",
                font=("Segoe UI", 13, "bold")).pack(side=tk.LEFT, padx=20, pady=15)
        
        fields_container = tk.Frame(selection_card, bg=COLOR_CARD)
        fields_container.pack(fill=tk.BOTH, padx=25, pady=25)
        
        fields = [
            ("Year", "📅 Academic Year"),
            ("Faculty", "🏛️ Faculty"),
            ("Regulation", "📜 Regulation"),
            ("Phase", "📊 Level (Ferka)"),
            ("Dept", "🎓 Department"),
            ("Semester", "📆 Semester"),
            ("Door", "🚪 Door"),
            ("SubjSem", "📚 Course Semester"),
            ("Subject", "📖 Course")
        ]
        
        for i, (key, label_text) in enumerate(fields):
            field_frame = tk.Frame(fields_container, bg=COLOR_CARD)
            field_frame.pack(fill=tk.X, pady=8)
            
            tk.Label(field_frame, text=label_text, bg=COLOR_CARD, fg=COLOR_TEXT,
                    font=("Segoe UI", 10, "bold"), anchor='w', width=20).pack(side=tk.LEFT)
            
            combo_wrapper = tk.Frame(field_frame, bg=COLOR_BORDER, bd=1)
            combo_wrapper.pack(side=tk.LEFT, fill=tk.X, expand=True)
            
            cb = ttk.Combobox(combo_wrapper, state="disabled", font=("Segoe UI", 10),
                            style="Modern.TCombobox")
            cb.pack(fill=tk.X, padx=1, pady=1)
            cb.bind("<<ComboboxSelected>>", lambda e, k=key: self.on_selection(k))
            self.combos[key] = cb
        
        options_card = tk.Frame(main_container, bg=COLOR_CARD)
        options_card.pack(fill=tk.X, pady=(0, 20))
        
        options_inner = tk.Frame(options_card, bg=COLOR_CARD)
        options_inner.pack(fill=tk.X, padx=25, pady=20)
        
        tk.Label(options_inner, text="⚙️ Export Options", bg=COLOR_CARD, fg=COLOR_TEXT,
                font=("Segoe UI", 12, "bold")).pack(anchor='w', pady=(0, 15))
        
        chk_frame = tk.Frame(options_inner, bg=COLOR_CARD)
        chk_frame.pack(anchor='w', fill=tk.X)

        self.chk_excel = tk.Checkbutton(
            chk_frame, text="📄 Save as Excel (.xlsx)", variable=self.var_save_excel,
            bg=COLOR_CARD, activebackground=COLOR_CARD, font=("Segoe UI", 10),
            fg=COLOR_TEXT, selectcolor=COLOR_CARD, cursor="hand2"
        )
        self.chk_excel.pack(side=tk.LEFT, padx=(0, 30))

        self.chk_pdf = tk.Checkbutton(
            chk_frame, text="📄 Save as PDF (requires Chrome)", variable=self.var_save_pdf,
            bg=COLOR_CARD, activebackground=COLOR_CARD, font=("Segoe UI", 10),
            fg=COLOR_TEXT, selectcolor=COLOR_CARD, cursor="hand2",
            command=self.on_toggle_pdf
        )
        self.chk_pdf.pack(side=tk.LEFT)
        
        chk_format_frame = tk.Frame(options_inner, bg=COLOR_CARD)
        chk_format_frame.pack(anchor='w', fill=tk.X, pady=(10,0))
        
        self.chk_format = tk.Checkbutton(
            chk_format_frame,
            text="🎨 Enable Visual Grades (Color Formatting for Excel)",
            variable=self.var_cond_format,
            bg=COLOR_CARD, activebackground=COLOR_CARD,
            font=("Segoe UI", 10), fg=COLOR_TEXT,
            selectcolor=COLOR_CARD, cursor="hand2"
        )
        self.chk_format.pack(side=tk.LEFT)
        
        progress_frame = tk.Frame(options_inner, bg=COLOR_CARD)
        progress_frame.pack(fill=tk.X, pady=(15, 0))
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate', length=300)
        self.progress.pack(fill=tk.X)
        
        actions_card = tk.Frame(main_container, bg=COLOR_CARD)
        actions_card.pack(fill=tk.X, pady=(0, 20))
        
        actions_inner = tk.Frame(actions_card, bg=COLOR_CARD)
        actions_inner.pack(fill=tk.X, padx=25, pady=25)
        
        tk.Label(actions_inner, text="🎯 Actions", bg=COLOR_CARD, fg=COLOR_TEXT,
                font=("Segoe UI", 12, "bold")).pack(anchor='w', pady=(0, 15))
        
        buttons_frame = tk.Frame(actions_inner, bg=COLOR_CARD)
        buttons_frame.pack(fill=tk.X)
        
        self.btn_cancel = tk.Button(
            buttons_frame, text="⏹ STOP", state="disabled",
            command=self.cancel_batch, bg=COLOR_WARNING, fg="white",
            activebackground="#dc2626", font=("Segoe UI", 10, "bold"),
            padx=20, pady=12, bd=0, cursor="hand2", relief=tk.FLAT, width=15
        )
        self.btn_cancel.pack(side=tk.LEFT, padx=(0, 10))
        
        self.btn_batch = tk.Button(
            buttons_frame, text="📚 DOWNLOAD ALL", state="disabled",
            command=self.confirm_batch_download, bg=COLOR_BATCH, fg="white",
            activebackground="#7c3aed", font=("Segoe UI", 10, "bold"),
            padx=20, pady=12, bd=0, cursor="hand2", relief=tk.FLAT, width=15
        )
        self.btn_batch.pack(side=tk.LEFT, padx=(0, 10))
        
        self.btn_fetch = tk.Button(
            buttons_frame, text="📥 DOWNLOAD CURRENT", state="disabled",
            command=self.start_single_fetch, bg=COLOR_SUCCESS, fg="white",
            activebackground="#059669", font=("Segoe UI", 10, "bold"),
            padx=20, pady=12, bd=0, cursor="hand2", relief=tk.FLAT, width=18
        )
        self.btn_fetch.pack(side=tk.RIGHT)
        
        log_card = tk.Frame(main_container, bg=COLOR_CARD)
        log_card.pack(fill=tk.BOTH, expand=True)
        
        log_header = tk.Frame(log_card, bg=COLOR_CARD)
        log_header.pack(fill=tk.X, padx=25, pady=15)
        
        tk.Label(log_header, text="📜 Activity Log", bg=COLOR_CARD, fg=COLOR_TEXT,
                font=("Segoe UI", 12, "bold")).pack(anchor='w')
        
        log_container = tk.Frame(log_card, bg=COLOR_CARD)
        log_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 25))
        
        self.log_area = scrolledtext.ScrolledText(
            log_container, height=12, state='disabled', font=("Consolas", 9),
            bg="#1e1e2e", fg="#a6e3a1", insertbackground="white",
            relief=tk.FLAT, bd=0, padx=10, pady=10
        )
        self.log_area.pack(fill=tk.BOTH, expand=True)
        
        for tag, color in [("info", "#a6e3a1"), ("error", "#f38ba8"), 
                          ("cyan", "#89dceb"), ("warning", "#fab387"), ("batch", "#cba6f7")]:
            self.log_area.tag_config(tag, foreground=color)
        
        tk.Frame(main_container.master, bg=COLOR_BG, height=30).grid(row=2, column=0)
        
        status_bar = tk.Frame(self.root, bg=COLOR_HEADER, height=35)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        status_bar.pack_propagate(False)
        
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(status_bar, textvariable=self.status_var, bg=COLOR_HEADER, fg="#9ca3af",
                font=("Segoe UI", 9), padx=20).pack(side=tk.LEFT, fill=tk.Y)

    def log(self, message, color="info"):
        def _log():
            self.log_area.config(state='normal')
            self.log_area.insert(tk.END, f"▸ {message}\n", color)
            self.log_area.see(tk.END)
            self.log_area.config(state='disabled')
            self.status_var.set(message)
        self.root.after(0, _log)

    def toggle_loading(self, is_loading):
        if is_loading:
            self.progress.start(10)
            self.root.config(cursor="watch")
        else:
            self.progress.stop()
            self.root.config(cursor="")

    def toggle_buttons(self, enable):
        state = "normal" if enable else "disabled"
        bg_fetch = COLOR_SUCCESS if enable else "#d1d5db"
        bg_batch = COLOR_BATCH if enable else "#d1d5db"
        
        if not self.is_batch_running:
            self.btn_fetch.config(state=state, bg=bg_fetch)
            self.btn_batch.config(state=state, bg=bg_batch)
        
        if self.is_batch_running:
            self.btn_cancel.config(state="normal")
        else:
            self.btn_cancel.config(state="disabled")

    def toggle_account(self):
        if self.is_batch_running:
            messagebox.showwarning("Please Wait", "Active download in progress. Cancel or wait for completion.")
            return
        
        self.current_profile = "Education" if self.current_profile == "Science" else "Science"
        self.btn_profile.config(text=ACCOUNTS[self.current_profile]['label'])
        
        self.log(f"🔄 Switching to {self.current_profile}...", "warning")
        self.reset_session()
        self.selections = {}
        self.dropdown_map = {}
        
        for cb in self.combos.values():
            cb.set('')
            cb['values'] = []
            cb['state'] = 'disabled'
        
        self.btn_fetch.config(state="disabled", bg="#d1d5db")
        self.btn_batch.config(state="disabled", bg="#d1d5db")
        
        threading.Thread(target=self.perform_login, daemon=True).start()

    def perform_login(self):
        try:
            self.toggle_loading(True)
            self.log(f"🔐 Authenticating as {self.current_profile}...", "cyan")
            
            creds = ACCOUNTS[self.current_profile]
            
            resp = self.session.get(LOGIN_URL, verify=False, timeout=30)
            soup = BeautifulSoup(resp.text, 'html.parser')
            data = self.get_form_data(soup)
            
            data.update({
                'txt_user_name': creds['user'],
                'txt_pw': creds['pass'],
                'Button1': 'دخول'
            })
            
            self.session.post(LOGIN_URL, data=data, headers={'Referer': LOGIN_URL}, 
                            verify=False, timeout=30)
            
            resp = self.session.get(TARGET_URL, verify=False, timeout=30)
            self.current_soup = BeautifulSoup(resp.text, 'html.parser')
            
            if "ddl_acad_year" not in resp.text:
                self.log("❌ Authentication failed", "error")
                self.toggle_loading(False)
                return
            
            self.log(f"✅ Connected as {self.current_profile}", "info")
            
            self.root.after(0, self.initiate_sequence)
        
        except Exception as e:
            self.log(f"❌ Connection error: {e}", "error")
            self.toggle_loading(False)

    def initiate_sequence(self):
        self.log("🚀 Auto-selecting latest year and faculty...", "cyan")
        try:
            year_opts = self.extract_options(self.id_map["Year"])[:4]
            if not year_opts:
                self.log("❌ Could not find any academic years to select.", "error")
                self.toggle_loading(False)
                return

            self.update_combo("Year", year_opts, auto_select_index=0)

        except Exception as e:
            self.log(f"❌ Auto-selection failed: {e}", "error")
            self.toggle_loading(False)

    def get_form_data(self, soup):
        data = {}
        if not soup: return data
        for inp in soup.find_all('input'):
            if inp.get('name'):
                data[inp.get('name')] = inp.get('value', '')
        return data

    def extract_options(self, element_id):
        if not self.current_soup: return []
        select = self.current_soup.find('select', {'id': element_id})
        if not select: return []
        
        opts = []
        for opt in select.find_all('option'):
            val = opt.get('value', '')
            txt = opt.text.strip()
            if val and val != '0' and "اختر" not in txt and not txt.startswith("Select"):
                opts.append((val, txt))
        return opts

    def update_combo(self, key, options, auto_select_index=None):
        cb = self.combos[key]
        cb['state'] = 'readonly'
        self.dropdown_map[key] = options
        cb['values'] = [txt for val, txt in options]
        
        self.toggle_loading(False)
        
        if options:
            if auto_select_index is not None and 0 <= auto_select_index < len(options):
                cb.current(auto_select_index)
                self.root.after(100, lambda: self.on_selection(key))
            else:
                cb.set(f"Select {key}...")
                cb.focus()
            
            if key == "Subject":
                self.btn_fetch.config(state="normal", bg=COLOR_SUCCESS)
                self.btn_batch.config(state="normal", bg=COLOR_BATCH)
        else:
            cb.set("No options found")
            cb['state'] = 'disabled'

    def make_post(self, updates, event_target):
        try:
            if not self.current_soup or not self.current_soup.find('input', {'name': '__VIEWSTATE'}):
                self.log("⚠️ Session expired. Please restart.", "warning")
                return False
            
            form_data = self.get_form_data(self.current_soup)
            
            for key, val in self.selections.items():
                html_id = self.id_map.get(key)
                if html_id:
                    elem = self.current_soup.find('select', {'id': html_id})
                    if elem and elem.get('name'):
                        form_data[elem.get('name')] = val
            
            for eid, val in updates.items():
                sel = self.current_soup.find('select', {'id': eid})
                if sel:
                    form_data[sel.get('name')] = val
            
            target_elem = self.current_soup.find('select', {'id': event_target})
            form_data['__EVENTTARGET'] = target_elem.get('name', '') if target_elem else ''
            
            if 'ctl00$ContentPlaceHolder3$Button1' in form_data:
                del form_data['ctl00$ContentPlaceHolder3$Button1']
            
            time.sleep(0.5)
            
            resp = self.session.post(TARGET_URL, data=form_data, 
                                    headers={'Referer': TARGET_URL}, 
                                    verify=False, timeout=45)
            
            if resp.status_code != 200:
                self.log(f"❌ Server error: {resp.status_code}", "error")
                return False
            
            self.current_soup = BeautifulSoup(resp.text, 'html.parser')
            return True
        
        except Exception as e:
            self.log(f"❌ Network error: {e}", "error")
            return False

    def load_step(self, step_name):
        try:
            get_opts = lambda k: self.extract_options(self.id_map[k])
            
            if step_name == "Year":
                opts = get_opts("Year")[:4]
                self.root.after(0, lambda: self.update_combo("Year", opts))
            elif step_name == "Faculty":
                opts = get_opts("Faculty")
                self.root.after(0, lambda: self.update_combo("Faculty", opts, auto_select_index=0))
            elif step_name == "Regulation":
                opts = get_opts("Regulation")
                self.root.after(0, lambda: self.update_combo("Regulation", opts))
            elif step_name == "Phase":
                opts = [o for o in get_opts("Phase") if "المستوى" in o[1]]
                self.root.after(0, lambda: self.update_combo("Phase", opts))
            elif step_name == "Dept":
                opts = get_opts("Dept")
                self.root.after(0, lambda: self.update_combo("Dept", opts))
            elif step_name == "Semester":
                opts = get_opts("Semester")
                self.root.after(0, lambda: self.update_combo("Semester", opts))
            elif step_name == "Door":
                opts = get_opts("Door")
                self.root.after(0, lambda: self.update_combo("Door", opts, auto_select_index=0))
            elif step_name == "SubjSem":
                opts = get_opts("SubjSem")
                self.root.after(0, lambda: self.update_combo("SubjSem", opts))
            elif step_name == "Subject":
                opts = get_opts("Subject")
                self.root.after(0, lambda: self.update_combo("Subject", opts))
        
        except Exception as e:
            self.log(f"❌ Error loading {step_name}: {e}", "error")
            self.toggle_loading(False)

    def on_selection(self, key):
        if self.is_batch_running: return
        
        idx = self.combos[key].current()
        if idx == -1: return
        
        val, text = self.dropdown_map[key][idx]
        self.selections[key] = val
        self.log(f"✓ Selected: {text}", "info")
        
        seq = ["Year", "Faculty", "Regulation", "Phase", "Dept", "Semester", "Door", "SubjSem", "Subject"]
        try:
            current_index = seq.index(key)
            next_steps = seq[current_index + 1:]
            for step in next_steps:
                if step in self.selections:
                    del self.selections[step]
                if step in self.combos:
                    self.combos[step].set('')
                    self.combos[step]['values'] = []
                    self.combos[step]['state'] = 'disabled'
                    self.btn_fetch.config(state="disabled", bg="#d1d5db")
                    self.btn_batch.config(state="disabled", bg="#d1d5db")
        except ValueError:
            pass
        
        self.toggle_loading(True)
        threading.Thread(target=self.process_next_step, args=(key, val)).start()

    def process_next_step(self, current_key, current_val):
        target_id = self.id_map.get(current_key)
        seq = ["Year", "Faculty", "Regulation", "Phase", "Dept", "Semester", "Door", "SubjSem", "Subject"]
        
        try:
            if current_key == "Subject":
                success = self.make_post({target_id: current_val}, target_id)
                if success:
                    self.log("✓ Ready to download", "cyan")
                self.toggle_loading(False)
                return
            
            if current_key == "Regulation":
                self.make_post({target_id: current_val}, target_id)
            elif current_key not in ["Year", "Faculty"]:
                self.make_post({target_id: current_val}, target_id)
            
            try:
                curr_idx = seq.index(current_key)
                next_step = seq[curr_idx + 1]
                self.load_step(next_step)
            except IndexError:
                self.toggle_loading(False)
        
        except Exception as e:
            self.log(f"❌ Processing error: {e}", "error")
            self.toggle_loading(False)

    def start_single_fetch(self):
        self.toggle_buttons(False)
        self.toggle_loading(True)
        threading.Thread(target=self.fetch_report_thread, args=(False,)).start()

    def confirm_batch_download(self):
        subjects = self.dropdown_map.get('Subject', [])
        
        if not subjects:
            messagebox.showerror("Error", "No subjects available to download.")
            return
        
        count = len(subjects)
        confirm = messagebox.askyesno(
            "Batch Download",
            f"Found {count} courses.\n\n"
            "This will download ALL of them automatically.\n"
            "You can cancel anytime.\n\n"
            "Start Download?"
        )
        
        if confirm:
            self.start_batch_process(subjects)

    def start_batch_process(self, subjects):
        self.is_batch_running = True
        self.stop_event.clear()
        
        self.btn_batch.config(state="disabled")
        self.btn_fetch.config(state="disabled")
        self.btn_cancel.config(state="normal", bg=COLOR_WARNING)
        self.toggle_loading(True)
        
        threading.Thread(target=self.batch_worker, args=(subjects,), daemon=True).start()

    def cancel_batch(self):
        if self.is_batch_running:
            self.stop_event.set()
            self.log("🛑 Cancelling...", "error")
            self.btn_cancel.config(text="Stopping...", state="disabled")

    def batch_worker(self, subjects):
        total = len(subjects)
        self.log(f"━━━ BATCH DOWNLOAD: {total} COURSES ━━━", "batch")
        
        subject_html_id = self.id_map["Subject"]
        
        for index, (val, text) in enumerate(subjects, 1):
            if self.stop_event.is_set():
                self.log("❌ Batch cancelled", "error")
                break
            
            self.log(f"[{index}/{total}] Processing: {text}", "batch")
            
            self.selections["Subject"] = val
            success = self.make_post({subject_html_id: val}, subject_html_id)
            if not success:
                self.log(f"❌ Failed: {text}", "error")
                continue
            
            self.fetch_report_thread(silent_save=True)
            time.sleep(1.0)
        
        self.is_batch_running = False
        self.stop_event.clear()
        self.root.after(0, lambda: self.finish_batch_ui())

    def finish_batch_ui(self):
        self.toggle_loading(False)
        self.btn_batch.config(state="normal")
        self.btn_fetch.config(state="normal")
        self.btn_cancel.config(text="⏹ STOP", state="disabled")
        self.log("━━━ BATCH COMPLETE ━━━", "batch")
        messagebox.showinfo("Complete", "Batch download finished successfully!")

    def fetch_report_thread(self, silent_save=False):
        try:
            self.log("📊 Fetching report...", "warning")
            form_data = self.get_form_data(self.current_soup)
            
            for key, val in self.selections.items():
                html_id = self.id_map.get(key)
                if html_id:
                    elem = self.current_soup.find('select', {'id': html_id})
                    if elem and elem.get('name'):
                        form_data[elem.get('name')] = val
            
            form_data['ctl00$ContentPlaceHolder3$Button1'] = 'تقرير الطلاب'
            
            resp = self.session.post(TARGET_URL, data=form_data, 
                                    headers={'Referer': TARGET_URL}, 
                                    verify=False, allow_redirects=False)
            
            if resp.status_code == 200:
                self.current_soup = BeautifulSoup(resp.text, 'html.parser')
            
            time.sleep(1.5)
            
            report_resp = self.session.get(REPORT_URL, headers={'Referer': TARGET_URL}, 
                                          verify=False)
            soup = BeautifulSoup(report_resp.text, 'html.parser')
            
            students_list, code, headers = self.parse_results(soup)
            
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            
            if students_list:
                self.log(f"✓ Found {len(students_list)} students", "cyan")
                
                if not self.var_save_excel.get() and not self.var_save_pdf.get():
                    self.log("⚠️ No output format selected. Nothing to save.", "warning")
                    if not silent_save:
                        messagebox.showwarning("No Selection", "Please select at least one output format (Excel or PDF) to download.")
                    return

                excel_path, pdf_path = None, None
                
                if self.var_save_excel.get():
                    excel_path = self.save_excel(students_list, code, headers, timestamp)
                
                if self.var_save_pdf.get():
                    pdf_path = self.save_chrome_pdf(report_resp.text, code, timestamp)

                if not silent_save:
                    saved_files = []
                    if excel_path: saved_files.append("Excel")
                    if pdf_path: saved_files.append("PDF")

                    if not saved_files:
                        messagebox.showerror("Save Failed", "Could not save any files. Please check the logs for errors.")
                        return

                    message = f"Saved { ' and '.join(saved_files) } successfully."
                    file_to_open = excel_path or pdf_path
                    
                    if file_to_open:
                        file_type = "Excel" if excel_path else "PDF"
                        message += f"\n\nOpen the {file_type} file now?"
                        if messagebox.askyesno("Success", message):
                            os.startfile(file_to_open)
                    else:
                        messagebox.showinfo("Success", message)
            else:
                self.log("⚠️ No students found", "warning")
                if not silent_save:
                    messagebox.showwarning("Result", "No students found.")
        
        except Exception as e:
            self.log(f"❌ Fetch error: {e}", "error")
        finally:
            if not self.is_batch_running:
                self.toggle_loading(False)
                self.toggle_buttons(True)

    def extract_table_headers(self, soup):
        header_row = soup.find('tr', align='center')
        if not header_row:
            return []
        
        cells = header_row.find_all('td')
        if len(cells) <= 6:
            return []
        
        target_cells = cells[:-6]
        
        extracted_headers = []
        for i, cell in enumerate(target_cells):
            text = cell.get_text(separator=" ", strip=True)
            if text in extracted_headers:
                text = f"{text}_{i}"
            extracted_headers.append(text)
        
        extracted_headers.reverse()
        return extracted_headers

    def parse_results(self, soup):
        txt = soup.get_text()
        match = re.search(r'([A-Za-z0-9]+)\s*كود المقرر', txt)
        code = match.group(1) if match else "Student_List"
        
        headers_list = self.extract_table_headers(soup)
        
        students = []
        table = soup.find('table', {'id': 'ctl00_ContentPlaceHolder3_gv_list'})
        rows = table.find_all('tr') if table else soup.find_all('tr')
        
        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 8:
                try:
                    s_txt = cells[-1].get_text(strip=True)
                    if not s_txt.isdigit(): continue
                    s_id = int(s_txt)
                    
                    seat_no = cells[-2].get_text(strip=True).replace('\xa0', '').strip()
                    name = cells[-3].get_text(strip=True).replace('\xa0', '').strip()
                    
                    grade_values = [c.get_text(strip=True) for c in cells[:-6]]
                    grade_values.reverse()
                    
                    if name and "اسم الطالب" not in name:
                        students.append({
                            'id': s_id,
                            'name': name,
                            'seat': seat_no,
                            'grades': grade_values
                        })
                except Exception:
                    continue
        
        students.sort(key=lambda x: x['id'])
        
        final_students = []
        seen_map = {}
        duplicates_removed = 0
        
        for s in students:
            name = s['name']
            s_id = s['id']
            
            if name not in seen_map:
                seen_map[name] = s
                final_students.append(s)
            else:
                duplicates_removed += 1
                original_student = seen_map[name]
                if original_student['id'] != s_id:
                    self.log(f"⚠️ Duplicate removed: {name}", "warning")
        
        if duplicates_removed > 0:
            self.log(f"✓ Cleaned {duplicates_removed} duplicates", "info")
        
        return final_students, code, headers_list

    def apply_conditional_formatting(self, ws, headers):
        self.log("🎨 Applying visual formatting...", "batch")
        
        red_error_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        orange_text_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
        
        for i, header_text in enumerate(headers):
            col_idx = i + 4
            col_letter = get_column_letter(col_idx)
            
            match = re.search(r"(\d+(\.\d+)?)", header_text)
            
            if match:
                max_val = float(match.group(1))
                if max_val == 0: max_val = 100
                mid_val = max_val / 2
                
                color_rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='F8696B',
                    mid_type='num', mid_value=mid_val, mid_color='FFEB84',
                    end_type='num', end_value=max_val, end_color='63BE7B'
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}1000", color_rule)
                
                error_rule = CellIsRule(operator='greaterThan', formula=[f"{max_val}"], 
                                       stopIfTrue=True, fill=red_error_fill)
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}1000", error_rule)
            
            text_rule = FormulaRule(formula=[f'ISTEXT({col_letter}2)'], 
                                   stopIfTrue=True, fill=orange_text_fill)
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}1000", text_rule)

    def save_excel(self, student_data, code, headers, timestamp):
        safe_code = "".join([c for c in code if c.isalnum() or c in (' ', '-', '_')]).strip()
        
        folder_name = "lists"
        if not os.path.exists(folder_name):
            try:
                os.makedirs(folder_name)
            except OSError as e:
                self.log(f"❌ Error creating folder: {e}", "error")
                return None
        
        filename = f"{safe_code}_{timestamp}.xlsx"
        file_path = os.path.join(folder_name, filename)
        
        data = []
        for i, s in enumerate(student_data, 1):
            row_dict = {
                'No': i,
                'Seat No': s['seat'],
                'Student Name': s['name']
            }
            
            s_grades = s.get('grades', [])
            for h_idx, header_title in enumerate(headers):
                if h_idx < len(s_grades):
                    val = s_grades[h_idx]
                    try:
                        row_dict[header_title] = float(val)
                    except ValueError:
                        row_dict[header_title] = val
                else:
                    row_dict[header_title] = ""
            
            data.append(row_dict)
        
        df = pd.DataFrame(data)
        
        standard_cols = ['No', 'Seat No', 'Student Name']
        final_cols = standard_cols + [h for h in headers if h in df.columns]
        existing_cols = [c for c in final_cols if c in df.columns]
        df = df[existing_cols]
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                ws = writer.sheets['Sheet1']
                ws.sheet_view.rightToLeft = True
                
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 45
                
                for i, _ in enumerate(headers):
                    try:
                        col_idx = i + 4
                        col_letter = get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = 22
                    except:
                        pass
                
                bold_font = Font(bold=True)
                for cell in ws['C']:
                    if cell.row > 1:
                        cell.font = bold_font
                
                if self.var_cond_format.get():
                    self.apply_conditional_formatting(ws, headers)
            
            self.log(f"✅ Excel Saved: {filename}", "info")
            return os.path.abspath(file_path)
        
        except PermissionError:
            self.log(f"❌ File is open: {filename}", "error")
            messagebox.showerror("Permission Error", 
                f"The file '{filename}' is currently open.\n\nPlease close it and try again.")
            return None
        except Exception as e:
            self.log(f"❌ Excel error: {e}", "error")
            return None

    # =========================================================
    # CHROME PDF GENERATION & VALIDATION
    # =========================================================

    def load_cached_chrome_path(self):
        """Checks for a saved config file in the lists folder."""
        try:
            config_path = os.path.join("lists", "chrome_config.json")
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    data = json.load(f)
                    # Check if 'ready' is explicitly True
                    if data.get('ready') is True:
                        path = data.get('chrome_path')
                        # Ensure the path on disk actually still exists
                        if path and os.path.exists(path):
                            return path
        except:
            pass
        return None

    def save_chrome_config(self, path):
        """Saves the verified Chrome path to lists/chrome_config.json."""
        try:
            folder = "lists"
            if not os.path.exists(folder):
                os.makedirs(folder)
            
            config_path = os.path.join(folder, "chrome_config.json")
            with open(config_path, 'w') as f:
                json.dump({"chrome_path": path, "ready": True}, f)
            self.log("💾 Config saved for future runs", "info")
        except Exception as e:
            self.log(f"⚠️ Could not save config: {e}", "warning")

    def find_local_chrome_path(self):
        """Attempts to find chrome.exe in common Windows locations."""
        
        # 1. Check Config File first
        cached = self.load_cached_chrome_path()
        if cached: return cached

        # 2. Check Standard Paths
        possible_paths = [
            os.path.join(os.environ.get('ProgramFiles', 'C:\\Program Files'), 'Google', 'Chrome', 'Application', 'chrome.exe'),
            os.path.join(os.environ.get('ProgramFiles(x86)', 'C:\\Program Files (x86)'), 'Google', 'Chrome', 'Application', 'chrome.exe'),
            os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Google', 'Chrome', 'Application', 'chrome.exe'),
            os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
        ]
        
        for path in possible_paths:
            if path and os.path.exists(path):
                return path
        return None

    def on_toggle_pdf(self):
        if self.var_save_pdf.get():
            self.toggle_loading(True)
            threading.Thread(target=self._check_chrome_thread, daemon=True).start()

    def _check_chrome_thread(self):
        is_chrome_ok = self.check_chrome_installation()
        def update_ui():
            self.toggle_loading(False)
            if not is_chrome_ok:
                self.var_save_pdf.set(False)
                messagebox.showerror(
                    "Chrome Not Found",
                    "Could not find a valid Google Chrome installation or its driver.\n\n"
                    "Please install Google Chrome to use the PDF export feature."
                )
        self.root.after(0, update_ui)

    def check_chrome_installation(self):
        # --- FAST PATH: Trust the config file if it exists and is valid ---
        cached_path = self.load_cached_chrome_path()
        if cached_path:
             self.log("⚡ Chrome verified from config.", "info")
             return True
        # ------------------------------------------------------------------

        # --- SLOW PATH: Run full Selenium check ---
        self.log("🔍 Verifying Chrome installation...", "cyan")
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--log-level=3")
            
            custom_path = self.find_local_chrome_path()
            if custom_path:
                self.log(f"📍 Found Chrome at: {custom_path}", "info")
                chrome_options.binary_location = custom_path
            else:
                self.log("⚠️ Could not auto-detect chrome.exe path, trying default...", "warning")

            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.quit()
            
            # Save config if successful and we have a path
            if custom_path:
                self.save_chrome_config(custom_path)
            
            self.log("✅ Chrome check successful.", "info")
            return True
        except Exception as e:
            self.log(f"❌ Chrome check failed. Error: {e}", "error")
            return False

    def save_chrome_pdf(self, html_content, code, timestamp):
        debug_mode = False

        folder_name = "lists"
        if not os.path.exists(folder_name):
            try:
                os.makedirs(folder_name)
            except OSError:
                pass
            
        safe_code = "".join([c for c in code if c.isalnum() or c in (' ', '-', '_')]).strip()
        filename = f"{safe_code}_{timestamp}.pdf"
        file_path = os.path.join(folder_name, filename)
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            if soup.head:
                soup.head.insert(0, soup.new_tag("base", href="https://tdb.tanta.edu.eg/student_results/"))
            
            for img in soup.find_all('img'):
                src = img.get('src')
                if not src: continue
                if not src.startswith("http"):
                    src = "https://tdb.tanta.edu.eg" + src if src.startswith("/") else src  
                try:
                    img_resp = self.session.get(src, headers={'Referer': REPORT_URL}, stream=True, timeout=10, verify=False)
                    if img_resp.status_code == 200:
                        encoded_string = base64.b64encode(img_resp.content).decode("utf-8")
                        ct = img_resp.headers.get('Content-Type', 'image/jpeg')
                        img['src'] = f"data:{ct};base64,{encoded_string}"
                except Exception: pass

            for table in soup.find_all('table', {'border': '1'}):
                table['class'] = table.get('class', []) + ['student-grade-table']
                for attr in ['style', 'border', 'cellspacing', 'cellpadding']:
                    if table.has_attr(attr): del table[attr]

            style = soup.new_tag('style')
            style.string = """
                @page { size: A4 portrait; margin: 0.5cm; }
                body { font-family: sans-serif; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
                table.student-grade-table { width: 100% !important; border-collapse: collapse !important; border-spacing: 0 !important; margin-top: 5px !important; }
                table.student-grade-table td, table.student-grade-table th { box-shadow: inset 0 0 0 0.5pt #000000 !important; border: none !important; padding: 3px !important; text-align: center; }
                table:not(.student-grade-table), table:not(.student-grade-table) td { border: none !important; }
            """
            if soup.head: soup.head.append(style)
            else: soup.body.insert(0, style)
            final_html = str(soup)
            
        except Exception as e:
            self.log(f"⚠️ HTML Prep failed: {e}", "error")
            final_html = html_content

        temp_html_path = os.path.abspath(f"temp_{timestamp}.html")
        try:
            with open(temp_html_path, "w", encoding="utf-8") as f:
                f.write(final_html)
                
            chrome_options = Options()
            if not debug_mode: chrome_options.add_argument("--headless")
            chrome_options.add_argument("--disable-gpu"); chrome_options.add_argument("--no-sandbox"); chrome_options.add_argument("--log-level=3")
            
            custom_path = self.find_local_chrome_path()
            if custom_path:
                chrome_options.binary_location = custom_path

            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            
            file_url = f"file:///{temp_html_path.replace(os.sep, '/')}"
            driver.get(file_url)
            time.sleep(1.5)
            
            print_options = {"landscape": False, "displayHeaderFooter": False, "printBackground": True, "preferCSSPageSize": False, "scale": 0.88, "paperWidth": 8.27, "paperHeight": 11.69}
            result = driver.execute_cdp_cmd("Page.printToPDF", print_options)
            
            with open(file_path, "wb") as f:
                f.write(base64.b64decode(result['data']))
                
            self.log(f"✅ PDF Saved: {filename}", "info")
            driver.quit()
            return os.path.abspath(file_path)
            
        except Exception as e:
            self.log(f"❌ Chrome Print Failed: {e}", "error")
            return None
        finally:
            if not debug_mode and os.path.exists(temp_html_path):
                try: os.remove(temp_html_path)
                except: pass

if __name__ == "__main__":
    root = tk.Tk()
    app = TantaScraperApp(root)
    root.mainloop()
